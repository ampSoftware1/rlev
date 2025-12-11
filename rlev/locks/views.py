import pytz
import json
import os
import uuid
import time
from urllib.parse import parse_qs
from django.db.models.functions import Length 

from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.decorators.http import require_POST
from django.db import IntegrityError, transaction

from django.db.models import Case, When, Value, BooleanField
from django.utils import timezone
from django.utils.timezone import localdate

from locks.phoneSys import *
from .login import LoginByApi
from .models import *
from django.views.decorators.csrf import csrf_exempt
# from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from .functions import *

class CustomLoginView(LoginView):
    authentication_backend = LoginByApi

    def get_success_url(self):
        next_url = self.request.GET.get('next')
        if next_url:
            view_to_run = next_url.replace('/', '')
        else:
            view_to_run = 'dashboard'
        
        return reverse_lazy(view_to_run)


def test(request):
    return render(request, 'locks/test.html')

@login_required
def dashboard(request):

    user = request.user
    is_admin = request.user.is_superuser
    amount_in_transmission = Transmission.get_amount_in_transmission()
    alerts_locks = Lock.get_alerts_electric_quantity()
    return render(request, 'locks/dashboard.html', {"user": user, 
                                                    "is_admin": is_admin, 
                                                    "amount_in_transmission": amount_in_transmission,
                                                    "alert_locks": alerts_locks})

def locks(request):

    is_admin = request.user.is_superuser
    if is_admin:
        locks_objects = Lock.objects.all()
    else:
        locks_objects = Lock.objects.filter(lockuser__user=request.user)
    
    locks_objects = locks_objects.order_by('lock_alias')
    locks_list = [lock_object.get_details() for lock_object in locks_objects]
    groups = AccessGroup.get_groups(request.user)

    return render(request, 'locks/locks.html', {"locks": locks_list,
                                                "groups": groups,
                                                "is_admin": is_admin})


def lock(request, lock_id):
    lock = Lock.objects.get(id=lock_id)
    is_admin = request.user.is_superuser
    user =  request.user
    lock_details = lock.get_details()
    return render(request, 'locks/lock.html', {"details": lock_details, "is_admin": is_admin, "user": user})


def get_lock_child_list(request, type_child, lock_id):
    lock = Lock.objects.get(id=lock_id)
    child_list = lock.get_child_list(type_child, request.user)
    return JsonResponse({'success': True, 'child_list': child_list})


def get_lock_link_list(request, type_link, lock_id):
    lock = Lock.objects.get(id=lock_id)
    link_list = lock.get_link_list(type_link)
    return JsonResponse({'success': True, 'link_list': link_list})


def get_lock_link_options(request, type_link, lock_id):
    if type_link == 'group':
        groups_object = AccessGroup.objects.exclude(lockaccessgroup__lock_id=lock_id)
        link_options = [{"id": group.id, "name": group.group_name} for group in groups_object]
    elif type_link == 'user':
        users_objects = User.objects.exclude(lockuser__lock_id=lock_id)
        link_options = [{"id": user.id, "name": f"{user.first_name} {user.last_name}"} for user in users_objects]
    return JsonResponse({'success': True, 'link_options': link_options})


def add_link_to_lock(request, type_link, lock_id, object_id):
    if type_link == 'group':
        new_object = LockAccessGroup(lock_id=lock_id,access_group_id=object_id)
    elif type_link == 'user':
        new_object = LockUser(lock_id=lock_id, user_id=object_id)
    new_object.save()

    return JsonResponse({'success': True}, status=200)


def remove_link_from_lock(request, type_link, object_id):
    if type_link == 'group':
        link_object = LockAccessGroup.objects.get(id=object_id)
    elif type_link == 'user':
        link_object = LockUser.objects.get(id=object_id)
    link_object.delete()

    return JsonResponse({'success': True}, status=200)


def delete_lock_child(request, type_child, object_id):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    if type_child == 'permissions':
        permissions_object = Permission.objects.get(id=object_id)
        permissions_object.delete_permission(user_name)

    return JsonResponse({'success': True}, status=200)


def check_lock_status(request, lock_id):
    lock = Lock.objects.get(id=lock_id)
    lock_id_ttl = lock.lock_id_ttl

    api = ApiRequest()
    status = api.get_lock_status(lock_id_ttl)

    if status:
        response = status
    else:
        response = {"success": False}

    return JsonResponse(response)


def change_lock_status(request, lock_id, current_status):
    lock = Lock.objects.get(id=lock_id)
    lock_id_ttl = lock.lock_id_ttl

    api = ApiRequest()
    if (current_status == 'locked'):
        change_lock = api.unlock(lock_id_ttl)
    else:
        change_lock = api.lock(lock_id_ttl)

    if change_lock:
        response = change_lock
    else:
        response = {"success": False}


    return JsonResponse(response)


def change_lock_alias(request, lock_id, new_alias):
    lock = Lock.objects.get(id=lock_id)
    lock_id_ttl = lock.lock_id_ttl
    lock.lock_alias = new_alias
    lock.save()

    api = ApiRequest()
    response = api.change_lock_alias(lock_id_ttl,new_alias)
    return JsonResponse(response)

def add_lock(request):
    return render(request, 'locks/add_lock.html')


def delete_lock(request, lock_id):
    lock = Lock.objects.get(id=lock_id)
    lock.delete()

    send_message_to_browser('refresh_locks')
    return JsonResponse({'success': True}, status=200)

def save_lock(request, lock_id_ttl):
    lock = Lock(lock_id_ttl=lock_id_ttl)
    lock.save()
    lock_id = lock.id
    return JsonResponse({'success': True, 'lock_id': lock_id}, status=200)


def check_lock(request, lock_id_ttl):
    api = ApiRequest()
    api_response = api.get_lock_details(lock_id_ttl)
    response = {}
    if api_response:
        lock_exists = Lock.objects.filter(lock_id_ttl=lock_id_ttl).exists()
        if lock_exists:
            response['success'] = False
            response['error'] = "מזהה המנעול כבר קיים במערכת"
        else:
            response['success'] = True
            response['lock_name'] = api_response['lockAlias']
    else:
        response['success'] = False
        response['error'] = "מזהה המנעול שהוזן אינו משותף לחשבון"
    return JsonResponse(response)

def select_locks(request):
    locks = Lock.get_locks(request.user)
    groups = AccessGroup.get_groups(request.user)

    return render(request, 'locks/select_locks.html', {"locks": locks, "groups": groups})


def passages(request):
    is_admin = request.user.is_superuser
    if is_admin:
        passages_objects = PassageMode.objects.all()
    else:
         passages_objects = PassageMode.objects.filter(passagemodeuser__user=request.user)

    passages = [passage.get_details() for passage in passages_objects]
    return render(request, 'locks/passages.html', {"passages": passages})


def passage(request, passage_id=None):
    is_admin = request.user.is_superuser

    data = {"is_admin": is_admin}
    if passage_id:
        data['type_view'] = 'edit' 
        passage = PassageMode.objects.get(id=passage_id)
        passage_details = passage.get_details()
        data['details'] = passage_details
    else:
        data['type_view'] = 'new' 

    days_in_week = [
    {"id": "1", "description": "ראשון"},  
    {"id": "2", "description": "שני"},
    {"id": "3", "description": "שלישי"},
    {"id": "4", "description": "רביעי"},
    {"id": "5", "description": "חמישי"},
    {"id": "6", "description": "שישי"},
    {"id": "7", "description": "שבת"}
    ]

    days_in_months = [
            {"id": "1", "description": "א"},
            {"id": "2", "description": "ב"},
            {"id": "3", "description": "ג"},
            {"id": "4", "description": "ד"},
            {"id": "5", "description": "ה"},
            {"id": "6", "description": "ו"},
            {"id": "7", "description": "ז"},
            {"id": "8", "description": "ח"},
            {"id": "9", "description": "ט"},
            {"id": "10", "description": "י"},
            {"id": "11", "description": "יא"},
            {"id": "12", "description": "יב"},
            {"id": "13", "description": "יג"},
            {"id": "14", "description": "יד"},
            {"id": "15", "description": "טו"},
            {"id": "16", "description": "טז"},
            {"id": "17", "description": "יז"},
            {"id": "18", "description": "יח"},
            {"id": "19", "description": "יט"},
            {"id": "20", "description": "כ"},
            {"id": "21", "description": "כא"},
            {"id": "22", "description": "כב"},
            {"id": "23", "description": "כג"},
            {"id": "24", "description": "כד"},
            {"id": "25", "description": "כה"},
            {"id": "26", "description": "כו"},
            {"id": "27", "description": "כז"},
            {"id": "28", "description": "כח"},
            {"id": "29", "description": "כט"},
            {"id": "30", "description": "ל"}
        ]

    months = [
        {"id": "7", "description": "תשרי"},
        {"id": "8", "description": "חשון"},
        {"id": "9", "description": "כסליו"},
        {"id": "10", "description": "טבת"},
        {"id": "11", "description": "שבט"},
        {"id": "12", "description": "אדר"},
        {"id": "13", "description": "אדר ב"},
        {"id": "1", "description": "ניסן"},  
        {"id": "2", "description": "אייר"},
        {"id": "3", "description": "סיון"},
        {"id": "4", "description": "תמוז"},
        {"id": "5", "description": "אב"},
        {"id": "6", "description": "אלול"},
    ]
    data['days_in_week'] = days_in_week
    data['days_in_months'] = days_in_months
    data['months'] = months
    # return JsonResponse({"passages": data})
    return render(request, 'locks/passage.html', data)

def get_passages_list(reqeust):
    passages_objects = PassageMode.objects.all()
    passages = [passage.get_details() for passage in passages_objects]
    return JsonResponse({"passages": passages})

def get_locks_passage_list(reqeust, passage_id):
    passage = PassageMode.objects.get(id=passage_id)
    locks_passage_objects = passage.passagemodelock_set.all()
    locks_passage = [lock.get_details() for lock in locks_passage_objects]
    return JsonResponse({"locks_passage": locks_passage})

def save_passage(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
        passage_id = data.pop('id', None)
        
        if passage_id:
            passage = PassageMode.objects.get(id=passage_id)
            for key, value in data.items():
                setattr(passage, key, value) 
        else:
            passage = PassageMode(**data)
            passage.send_webhook = 1

        passage.done = 0
        passage.save()

        user_add = request.user
        if not user_add.is_superuser and not passage_id:
            passage_user = PassageModeUser(passage_mode=passage, user=user_add)
            passage_user.save()


        return JsonResponse({'success': True, "id": passage.id}, status=200)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


def delete_passage(request, passage_id):
    passage = PassageMode.objects.get(id=passage_id)
    active_passage_locks = passage.passagemodelock_set.filter(passage_status=1).exists()
    if active_passage_locks:
        return JsonResponse({"success": False, 'error': "לא ניתן למחוק מצב מעבר כשיש מנעולים עם מצב מעבר פעיל"})

    passage.delete()
    return JsonResponse({"success": True})

def change_active_passage(request, passage_id, value):
    passage = PassageMode.objects.get(id=passage_id)

    passage.active = value
    passage.save()
    return JsonResponse({"success": True})

def add_locks_passage(request, passage_id):
    locks_ids = json.loads(request.body)
    for lock in locks_ids: 
        lock_passage, create = PassageModeLock.objects.get_or_create(lock_id=lock, passage_mode_id=passage_id)
        lock_passage.passage_status = 2
        lock_passage.save()


    passage_mode = PassageMode.objects.get(id=passage_id)
    passage_mode.change_modes()
    
    return JsonResponse({"success": True})

def do_passage_locks_actions(request, passage_id):
    passage_mode = PassageMode.objects.get(id=passage_id)
    passage_mode.do_locks_actions()
    return JsonResponse({"success": True})

def get_passage_users(request, passage_id):
    passage = PassageMode.objects.get(id=passage_id)
    users = passage.passagemodeuser_set.all()
    users_list = [{"id": user.id, "name": user.get_name()} for user in users]
    return JsonResponse({'success': True, 'users_list': users_list})

def remove_user_from_passage(request,user_id):
    PassageModeUser.objects.get(id=user_id).delete()
    return JsonResponse({'success': True}, status=200)

def get_user_passage_options(request, passage_id):
    users_objects = User.objects.exclude(passagemodeuser__passage_mode_id=passage_id)
    users_options = [{"id": user.id, "name": f"{user.first_name} {user.last_name}"} for user in users_objects]
    return JsonResponse({'success': True, 'users_options': users_options})

def add_user_to_passage(request, passage_id, user_id):
    new_user = PassageModeUser(passage_mode_id=passage_id, user_id=user_id)
    new_user.save()

    return JsonResponse({'success': True}, status=200)


def lock_passage_action(request, action, lock_passage_id):
    lock_passage = PassageModeLock.objects.get(id=lock_passage_id)
    if action == 'delete':
        lock_passage.delete()
    else:
        current_time = int(time.time())
        if action == 'reset_passage':
            lock_passage.passage_time = current_time
        elif action == 'reset_unlock':
            lock_passage.unlock_time = current_time
        elif action == 'cancel':
            lock_passage.mode = 3
            lock_passage.passage_status = 0
            lock_passage.passage_time = current_time
        elif action == 'restart':
            lock_passage.mode = 1
            lock_passage.passage_status = 0
            lock_passage.passage_time = current_time

        lock_passage.save()
        lock_passage.change_passage_mode()
        

    return JsonResponse({"success": True})

def houses(request):
    is_admin = request.user.is_superuser
    if is_admin:
        houses_objects = House.objects.all()
    else:
        houses_objects = House.objects.filter(houseuser__user=request.user)

    houses_objects = houses_objects.order_by('description')
   
    houses = [house.get_details() for house in houses_objects]

    return render(request, 'locks/houses.html', {"houses": houses, 'locks': locks, 'is_admin': is_admin})


def houses_reports(request):
    return render(request, 'locks/houses_reports.html')

def get_houses_report_old(request,type_get, value):
    
    if type_get == 'date' or type_get == 'new':
        date = value
        check_exist_report = HousesReport.objects.filter(report_date=date)
        if type_get == 'date' and check_exist_report.exists():
            report = check_exist_report.last()
        else:
            report = HousesReport.create_new_report(date)
            
            houses = House.objects.filter(active=1)
            if not request.user.is_superuser:
                houses = houses.filter(houseuser__user=request.user)
            house_with_checkout_in_date = list(Hosting.objects.filter(lodging_end=date).values_list('house_id', flat=True))

            records = []
            for house in houses:
                record = HousesReportRecord(report=report, house=house)
                if house.id in house_with_checkout_in_date:
                    record.clean_check_out = 1
                records.append(record)

            HousesReportRecord.objects.bulk_create(records)
    elif type_get == 'id':
        id = value
        report = HousesReport.objects.get(id=id)
    elif type_get == 'duplicate':
        values = value.split(";")
        report_id = values[0]
        to_date = values[1]

        report = HousesReport.create_new_report(to_date)
        current_record = HousesReportRecord.objects.filter(report_id=report_id)
    
        with transaction.atomic():
            for record in current_record:
                record.pk = None 
                record.report = report
                record.save()


    houses_records = report.get_houses_records()
    html_houses_records = render_to_string('locks/houses_report_list.html',{'houses_records': houses_records}) 

    list_reports = HousesReport.get_list_reports()

    details =dict(
        id=report.id,
        report_description=report.description,
        report_date=report.report_date.strftime('%Y-%m-%d'),
        html_houses_records=html_houses_records,
        list_reports=list_reports
    )
    return JsonResponse(details)

def update_houses_report(request, report_id):
    report = HousesReport.objects.get(id=report_id)

    data = json.loads(request.body)

    report_description = data['report_description']
    report.description = report_description
    report.save()

    updates = []

    for record_id, fields in data['houses_records'].items():
        updates.append(
            HousesReportRecord(
                id=record_id,
                clean=fields.get("clean", False),
                clean_check_out=fields.get("clean_check_out", False),
                meal=fields.get("meal", False),
                fault=fields.get("fault", False),
                bed=fields.get("bed", False),
                note=fields.get("note", '')
            )
        )

    HousesReportRecord.objects.bulk_update(
        updates,
        ['clean', 'clean_check_out', 'meal', 'fault', 'bed', 'note']
    )

    return JsonResponse({"success": True})


def delete_houses_report(request, report_id):
    HousesReport.objects.get(id=report_id).delete()
    return JsonResponse({"success": True})

def get_houses_report_pdf(request):
    import pdfkit
    data = json.loads(request.body)
    report_id = data['report_id']
    houses_records_ids = data['houses_records_ids']
    type_report = data['type_report']

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{report_id}.pdf"'

    report = HousesReport.objects.get(id=report_id)
    houses_records = report.get_houses_records(houses_records_ids)
    report_details = dict(
        houses_records=houses_records,
        report_date=report.report_date,
        report_description=report.description   
    )
    html_content = render_to_string('locks/houses_report_' + type_report + '_pdf.html', report_details)

    options = {
        'page-size': 'A4',
        'encoding': 'UTF-8',
    }

    if type_report == 'details':
        options['orientation'] = 'Landscape'

    pdf_content = pdfkit.from_string(html_content, False, options=options)

    response.write(pdf_content)
    return response

def get_hostings_houses_for_date(request, date):
    if len(date) > 3:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    else:
        date_obj = None

    is_admin = request.user.is_superuser
    if is_admin:
        houses_objects = House.objects.all()
    else:
        houses_objects = House.objects.filter(houseuser__user=request.user)
    
    houses = {}
    for house in houses_objects:
        houses[house.id] = house.get_hostings_by_date(date_obj)
        
    
    return JsonResponse({"houses": houses})

##############################################

def select_houses(request):
    houses = House.get_houses(request.user)

    return render(request, 'locks/select_houses.html', {"houses": houses})



def get_houses_report(request,type_get, value):
    
    if type_get == 'date' or type_get == 'new':
        date = value
        check_exist_report = HousesReport.objects.filter(report_date=date)
        if type_get == 'date' and check_exist_report.exists():
            report = check_exist_report.last()
        else:
            report = HousesReport.create_new_report(date)
            
            houses = House.objects.filter(active=1)
            if not request.user.is_superuser:
                houses = houses.filter(houseuser__user=request.user)
            house_with_checkout_in_date = list(Hosting.objects.filter(lodging_end=date).values_list('house_id', flat=True))

            records = []
            for house in houses:
                record = HousesReportRecord(report=report, house=house)
                if house.id in house_with_checkout_in_date:
                    record.clean_check_out = 1
                records.append(record)

            HousesReportRecord.objects.bulk_create(records)
    elif type_get == 'id':
        id = value
        report = HousesReport.objects.get(id=id)
    elif type_get == 'duplicate':
        values = value.split(";")
        report_id = values[0]
        to_date = values[1]

        report = HousesReport.create_new_report(to_date)
        current_record = HousesReportRecord.objects.filter(report_id=report_id)
    
        with transaction.atomic():
            for record in current_record:
                record.pk = None 
                record.report = report
                record.save()


    houses_records = report.get_houses_records()
    html_houses_records = render_to_string('locks/houses_report_list.html',{'houses_records': houses_records}) 

    list_reports = HousesReport.get_list_reports()

    details =dict(
        id=report.id,
        report_description=report.description,
        report_date=report.report_date.strftime('%Y-%m-%d'),
        html_houses_records=html_houses_records,
        list_reports=list_reports
    )
    return JsonResponse(details)

#########################

def house(request, house_id=None):
    data = {}
    locks = Lock.get_locks()
    data['locks'] = locks
    data['is_admin'] = request.user.is_superuser
    if house_id:
        house_object = House.objects.get(id=house_id)
        data['details'] = house_object.get_details()
        data['type_view'] = 'edit'
    else:
        data['type_view'] = 'new'

    return render(request, 'locks/house.html', data)


def add_locks_to_house(request, house_id):
    locks_ids = json.loads(request.body)
    for lock in locks_ids: 
        lock_house, create = LocksHouse.objects.get_or_create(house_id=house_id,lock_id=lock)
        lock_house.save()

    return JsonResponse({"success": True})


def get_house_link_list(request, type_link, house_id):
    house = House.objects.get(id=house_id)
    link_list = house.get_link_list(type_link)
    return JsonResponse({'success': True, 'link_list': link_list})

def get_house_link_options(request, type_link, house_id):
    if type_link == 'user':
        users_objects = User.objects.exclude(houseuser__house_id=house_id)
        link_options = [{"id": user.id, "name": f"{user.first_name} {user.last_name}"} for user in users_objects]
    return JsonResponse({'success': True, 'link_options': link_options})


def add_link_to_house(request, type_link, house_id, object_id):
    if type_link == 'user':
        new_object = HouseUser(house_id=house_id, user_id=object_id)
    new_object.save()

    return JsonResponse({'success': True}, status=200)

def remove_link_from_house(request, type_link, object_id):
    if type_link == 'lock':
        link_object = LocksHouse.objects.get(id=object_id)
    elif type_link == 'user':
        link_object = HouseUser.objects.get(id=object_id)
    link_object.delete()

    return JsonResponse({'success': True}, status=200)


def get_locks_house_list(request, house_id):
    house = House.objects.get(id=house_id)
    locks_house_objects = house.lockshouse_set.all()
    locks_house = [{"id": lock_house.id, "name": lock_house.lock.lock_alias} for lock_house in locks_house_objects]
    return JsonResponse({"locks_house": locks_house})


def get_house_child_list(request, type_child, house_id):
    house = House.objects.get(id=house_id)
    child_list = house.get_child_list(type_child)
    return JsonResponse({'success': True, 'child_list': child_list})

def save_house(request):
    house_data = json.loads(request.body.decode('utf-8'))
    house_id = house_data['house_id']
    if len(house_id):
        house = House.objects.get(id=house_id)
        for key, value in house_data.items():
            setattr(house, key, value)
    else:
        house_data.pop("house_id")
        house = House(**house_data)
   
    house.save()
    house_id = house.id
    
    return JsonResponse({"success": True, "house_id": house_id})
    
    
def delete_house(request, house_id):
    House.objects.get(id=house_id).delete()
    return JsonResponse({"success": True})

def get_houses_with_hosting_status(request):
    hosting_data = json.loads(request.body.decode('utf-8'))
    lodging_start = hosting_data['lodging_start']
    lodging_end = hosting_data['lodging_end']
    by_hosting_id = hosting_data['by_hosting_id']
    houses = House.get_houses_with_hosting_status(lodging_start, lodging_end, by_hosting_id)
    return JsonResponse({"houses": houses})

def houses_users_permissions(request):
    users_objects = User.objects.all()
    users_details = [{"id": user.id,
                "description": user.first_name + ' ' + user.last_name,
                "count_houses": HouseUser.objects.filter(user_id=user.id).count()}
                for user in users_objects]
    return render(request, 'locks/houses_users_permissions.html', {"users": users_details})


def get_houses_users_permissions(request, user_id):
    child_houses = HouseUser.get_user_permission(user_id)
    return JsonResponse(child_houses, status=200)

def remove_house_from_user(request, user_id, house_id):
    house_user = HouseUser.objects.get(user_id=user_id, house_id=house_id)
    house_user.delete()
    child_houses = HouseUser.get_user_permission(user_id)
    return JsonResponse(child_houses, status=200)

def add_houses_to_user(request):
    data = json.loads(request.body)

    user_id = data['user_id']
    
    houses_ids = data['houses']
    for house in houses_ids:
        HouseUser.objects.update_or_create(user_id=user_id, house_id=house)

    child_houses = HouseUser.get_user_permission(user_id)

    return JsonResponse(child_houses, status=200)


def hostings(request):
    wards = Hosting.get_all_hospital_wards()
    houses = House.get_houses(request.user)
    return render(request, 'locks/hostings.html', {"wards": wards, 
                                                   "houses": houses})


def hosting(request, hosting_id=None):
    data = {}
    houses = House.get_houses(request.user)
    data['houses'] = houses
    if hosting_id:
        hosting_object = Hosting.objects.get(id=hosting_id)
        if not request.user.is_superuser:
            house = hosting_object.house
            check_exist = HouseUser.objects.filter(house=house, user=request.user).exists()
            if not check_exist:
                return render(request, 'locks/hosting_not_permission.html')

        data['details'] = hosting_object.get_details()
        data['type_view'] = 'edit'
    else:
        data['type_view'] = 'new'

    return render(request, 'locks/hosting.html', data)

def doplicate_hosting(request, hosting_id):
    data = {}
    houses = House.get_houses(request.user)
    data['houses'] = houses
    
    hosting_object = Hosting.objects.get(id=hosting_id)
    source_details = hosting_object.get_details()

    keys_to_remove = ['id','documents', 'file_path1', 'file_path2', 'note', 'lodging_start', 'lodging_end']
    new_details = {key: value for key, value in source_details.items() if key not in keys_to_remove}

    data['details'] = new_details
    data['type_view'] = 'new'

    return render(request, 'locks/hosting.html', data)

def save_hosting(request):
    hosting_data = json.loads(request.body.decode('utf-8'))
    hosting_id = hosting_data['hosting_id']
    if hosting_id:
        hosting = Hosting.objects.get(id=hosting_id)
    else:
        hosting = Hosting()
    
    hosting.guest = get_person_by_details(hosting_data['guest_details'])

    guest_is_patient = hosting_data['guest_is_patient']
    hosting.guest_is_patient = guest_is_patient
    if guest_is_patient:
        hosting.patient = None
    else:
        hosting.patient = get_person_by_details(hosting_data['patient_details'])

    house = House.objects.get(id=hosting_data['house_id'])
    hosting.house = house
   
    lodging_start = hosting_data['lodging_start']
    lodging_end = hosting_data['lodging_end']
    
    hosting.hospital_ward = hosting_data['hospital_ward']
    if not hosting_id and len(hosting_data['hospital_ward']) > 0:
        from .translite_ai import translate_hospital_department 
        hosting.hospital_ward_eng = translate_hospital_department(hosting_data['hospital_ward'])

    hosting.lodging_start = lodging_start
    hosting.lodging_end = lodging_end
    hosting.affinity = hosting_data['affinity']
    hosting.documents = hosting_data['documents']
    hosting.trigger = hosting_data['trigger']
    hosting.note = hosting_data['note']
    hosting.file_path1 = hosting_data['file_path1']
    hosting.file_path2 = hosting_data['file_path2']
    hosting.persons_in_house = hosting_data['persons_in_house']

    hosting.save()
    hosting_id = hosting.id
    
    return JsonResponse({"success": True, "hosting_id": hosting_id})
    

def get_person_by_details(details):
    id_number = details['id_number']
    person, create = Person.objects.get_or_create(id_number=id_number)
    for key, value in details.items():
        setattr(person, key, value)
    if create:
        from .translite_ai import transliterate_hebrew_name
        if len(person.first_name) > 0:  
            person.first_name_eng = transliterate_hebrew_name(person.first_name)
        if len(person.last_name) > 0 :
            person.last_name_eng = transliterate_hebrew_name(person.last_name)
    person.save()
    return person

def delete_hosting(request, hosting_id):
    Hosting.objects.get(id=hosting_id).delete()
    return JsonResponse({"success": True})
   

def get_hosting_record(request, guest_id_number):
    guest = Person.objects.get(id_number=guest_id_number)

    record = None
    hostings = None

    record_check = HostingRecord.objects.filter(guest=guest)
    if record_check.exists():
        record = record_check.last().get_details()
    
    hostings_check = Hosting.objects.filter(guest=guest)
    hostings = [hosting.get_details() for hosting in hostings_check]
    
    return JsonResponse({"record":record, "hostings": hostings})
    
def get_details_of_guest(request, guest_id_number):
    guest = Person.objects.get(id_number=guest_id_number)
    details = guest.get_details()
    return JsonResponse(details)
    

def remove_lock_from_house(request, lock_house_id):
    LocksHouse.objects.get(id=lock_house_id).delete()
    return JsonResponse({"success": True})

def permissions(request):   
    roles = Person.objects.values_list('role', flat=True).distinct()
    return render(request, 'locks/permissions.html', {"roles": roles})

def permission(request, permission_id=None):
    permission = Permission.objects.get(id=permission_id)
    persons = Person.get_persons()
    permission_details = permission.get_details()
   
    if permission.type_object_id == 8:
        card = Card.objects.get(id=permission.content_object.id)
        person_id = card.person.id
        cards_ids = "[" + str(card.id) + "]"
        phones_ids = '[]'
    else:
        phone = Phone.objects.get(id=permission.content_object.id)
        person_id = phone.person.id
        cards_ids = '[]'
        phones_ids = "[" + str(phone.id) + "]"

    data = {"details": permission_details,
            "person_id": person_id,
            "locks_ids": "[" + str(permission.lock.id) + "]",
            "cards_ids": cards_ids,
            "phones_ids": phones_ids,
            "persons": persons, 
            "type_view": 'edit'}

    # return JsonResponse(data)
    return render(request, 'locks/permission.html', data)

def get_hosting_options(request):
    triggers = Hosting.get_all_triggers()
    hospital_wards = Hosting.get_all_hospital_wards()
    triggers = [trigger for trigger in triggers]
    hospital_wards = [hospital_ward for hospital_ward in hospital_wards]
    return JsonResponse({"triggers": triggers, "hospital_wards": hospital_wards})


def add_permissions(request):
    add_permissions_data = json.loads(request.body.decode('utf-8'))

    by_object = add_permissions_data['by_object']
    persons = Person.get_persons()

    data = dict(
        type_view='new',
        by_object=by_object,
        persons=persons,
        cards_ids="[]",
        locks_ids="[]",
        locks_descriptions="[]",
        phones_ids="[]",
    )

    if by_object == 'multy':
        cards_ids = add_permissions_data.get('cards', [])
        if isinstance(cards_ids, str):
            cards = json.loads(cards_ids)
        else:
            cards = cards_ids
        phones_ids = add_permissions_data.get('phones', [])
        
        data['cards_ids'] = cards_ids
        data['phones_ids'] = phones_ids

        from_hosting = add_permissions_data.get('from_hosting', False)
        if from_hosting:
            house_id = add_permissions_data['house_id']
            house = House.objects.get(id=house_id)
            locks_house = house.lockshouse_set.all()
            locks_ids = []
            locks_descriptions = []

            if house.lock:
                locks_ids.append(house.lock.id)
                locks_descriptions.append(house.lock.lock_alias)

            for lock_house in locks_house:
                locks_ids.append(lock_house.lock.id)
                locks_descriptions.append(lock_house.lock.lock_alias)
            data['locks_ids'] = locks_ids
            data['locks_descriptions'] = json.dumps(locks_descriptions, ensure_ascii=False)

            lodging_start = add_permissions_data['lodging_start']
            lodging_end = add_permissions_data['lodging_end']
            
            lodging_start_time = int(datetime.strptime(lodging_start, "%Y-%m-%d").timestamp()) * 1000
            lodging_end_time = int(datetime.strptime(lodging_end, "%Y-%m-%d").timestamp()) * 1000
            current_time = int(datetime.now().timestamp()) * 1000

            data['details'] = {}
            type_permission = 2
            start_date = lodging_start_time
            end_date = lodging_end_time + (21 * 60 * 60) * 1000

            exists_permissions = Permission.objects.filter(type_object_id=8, object_id__in=cards, lock_id__in=locks_ids, type_permission__in=[1,2])
            if exists_permissions.exists():
                has_constant_permission = exists_permissions.filter(type_permission=1).exists()
                if has_constant_permission:
                    type_permission = 1
                else: 
                    max_end_date = exists_permissions.order_by('-end_date').first().end_date
                    if max_end_date > lodging_end_time:
                        end_date = max_end_date
                
                    if lodging_start_time > current_time:
                        active_permission = exists_permissions.filter(start_date__lt=lodging_start_time,end_date__gt=current_time).order_by('start_date')
                        if active_permission.exists():
                            start_date = active_permission.first().start_date
            
            data['details']['type_permission'] = type_permission
            if type_permission == 2:
                data['details']['start_date'] = start_date
                data['details']['end_date'] = end_date

    elif by_object == 'lock':
        lock = Lock.objects.get(id=add_permissions_data['lock_id'])
        data["locks_ids"] = "[" + str(add_permissions_data['lock_id']) + "]"
        data["lock_alias"] = lock.lock_alias

    elif by_object == 'person':
        data["person_id"] = int(add_permissions_data['person_id'])
        if 'card_id' in add_permissions_data:
            card_id = add_permissions_data['card_id']
            data["cards_ids"] = "[" + str(card_id) + "]"
            data["card_description"] = Card.objects.get(id=card_id).card_name
        

    return render(request, 'locks/permission.html', data)
    # return JsonResponse(data, status=200)


def save_permissions(request):
    # try:
    data = json.loads(request.body.decode('utf-8'))
    locks = data.get("locks", [])
    cards = data.get("cards", [])
    phones = data.get("phones", [])

    permission_data = data["permission_data"]
    permission_data['grant_by'] = request.user

    for lock in locks:
        lock_id = lock
        permission_data['lock_id'] = lock_id

        for card in cards:
            card_id = card
            Permission.objects.filter(object_id=card_id, type_object_id=8, lock_id=lock_id, status_record=1).delete()
            Permission.objects.filter(object_id=card_id, type_object_id=8, lock_id=lock_id, status_record=0).update(status_record=3)

            permission_data["type_object_id"] = 8
            permission_data["object_id"] = card_id
            permission_data['status_record'] = 1

            new_permission = Permission(**permission_data)
            new_permission.save()

            transmission_record = Transmission()
            transmission_record.type_action = 1
            transmission_record.lock_name = new_permission.lock.lock_alias
            transmission_record.card_number = new_permission.content_object.card_number
            transmission_record.card_description = new_permission.content_object.get_full_description()
            transmission_record.permission_id = new_permission.id
            transmission_record.user_transmission = f"{request.user.first_name} {request.user.last_name}"
            transmission_record.save()
                

        for phone in phones:
            phone_id = phone
            Permission.objects.filter(object_id=phone_id, type_object_id=16, lock_id=lock_id).delete()

            permission_data["type_object_id"] = 16
            permission_data["object_id"] = phone_id
            permission_data['status_record'] = 0

            new_permission = Permission(**permission_data)
            new_permission.save()

    return JsonResponse({'success': True}, status=200)

    # except Exception as e:
    #     return JsonResponse({'success': False, 'error': str(e)}, status=400)


def get_amount_in_transmission(request):
    amount_in_transmission = Transmission.get_amount_in_transmission()
    return JsonResponse({'amount_in_transmission': amount_in_transmission}, status=200)


def remove_permissions(request):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    data = json.loads(request.body.decode('utf-8'))
    
    permissions = data.get("permissions", [])
    if permissions:
        permissions_to_delete = Permission.objects.filter(id__in=permissions)
    else:
        person_id = data.get("person_id", False)
        
        if person_id:
            cards = Card.objects.filter(person_id=person_id).values_list('id', flat=True)
            phones = Phone.objects.filter(person_id=person_id).values_list('id', flat=True)
        else:

            cards = data.get("cards", [])
            phones = data.get("phones", [])

        permissions_to_delete = Permission.objects.filter(
            Q(type_object_id=8, object_id__in=cards) | Q(type_object_id=16, object_id__in=phones)
        )

    locks_to_sync = list(set(permission.lock.id for permission in permissions_to_delete if permission.type_object_id == 8))

    for permission in permissions_to_delete:
        permission.delete_permission(user_name)
    
    return JsonResponse({'success': True, 'locks_to_sync': locks_to_sync}, status=200)



def cards(request):
    return render(request, 'locks/cards.html')


def childs_locks(request, child_name):
    if child_name == 'group':
        child_description = 'קבוצות גישה'
        childs = AccessGroup.get_groups()
    elif child_name == 'user':
        child_description = 'הרשאות משתמשים'
       
        users_objects = User.objects.all()
        childs = [{"id": user.id,
                   "description": user.first_name + ' ' + user.last_name,
                   "count_locks": LockUser.objects.filter(user_id=user.id).count()}
                  for user in users_objects]
    return render(request, 'locks/childs_locks.html', {"child_name": child_name, "child_description": child_description, "childs": childs})


def access_group(request):
    return render(request, 'locks/access_group.html')


def add_access_group(request, group_name):
    new_group = AccessGroup(group_name=group_name)
    new_group.save()
    return JsonResponse({"success": True})


def delete_group(request, group_id):
    AccessGroup.objects.get(id=group_id).delete()
    return JsonResponse({"success": True})


def get_child_locks(request, child_name, child_id):
    child_locks = {}
    if child_name == 'group':
        child_locks = LockAccessGroup.get_group_locks(child_id)
    elif child_name == 'user':
        child_locks = LockUser.get_user_locks(child_id)
    return JsonResponse(child_locks, status=200)


def add_or_remove_locks_to_child(request):
    data = json.loads(request.body)

    action = data['action']
    child_id = data['child_id']
    child_name = data['child_name']

    if action == 'add':
        locks_ids = data['locks']
        for lock in locks_ids:
            if child_name == 'group':
                LockAccessGroup.objects.update_or_create(access_group_id=child_id,lock_id=lock)
            elif child_name == 'user':
                LockUser.objects.update_or_create(user_id=child_id, lock_id=lock)

    elif action == 'remove':
        lock_id = data['lock_id']
        if child_name == 'group':
            LockAccessGroup.objects.get(access_group_id=child_id, lock_id=lock_id).delete()
        elif child_name == 'user':
            LockUser.objects.get(user_id=child_id, lock_id=lock_id).delete()

    child_locks = {}
    if child_name == 'group':
        child_locks = LockAccessGroup.get_group_locks(child_id)
    elif child_name == 'user':
        child_locks = LockUser.get_user_locks(child_id)

    return JsonResponse(child_locks, status=200)


def persons(request):
    is_admin = request.user.is_superuser
    roles = Person.get_all_roles()
    roles = [role for role in roles]
    return render(request, 'locks/persons.html', {"persons": persons, "is_admin": is_admin, "roles": roles})

def persons_list(request):
    details = {}
    data = json.loads(request.body)

    is_admin = request.user.is_superuser
    allow_feedback = is_admin or request.user.allow_feedback
    end_index = 0
    if not "role" in data and not "search_person" in data: 
        current_index = data.get("current_index", 0)
        end_index = int(current_index) + 50
        persons_objects = Person.objects.annotate(phone_length=Length('person_phone')).filter(phone_length__gt=0, not_to_view=False).order_by('-date_add')[current_index:end_index]
    else:
        persons_objects = Person.objects
        if "role" in data:
            persons_objects = persons_objects.filter(role=data['role'])
        if "search_person" in data:
            value = data['search_person']
            
            if value.isdigit():
                reverse_value = reverse_card_number(value)
                cards = Card.objects.filter(Q(card_number__icontains=value) | Q(card_number=reverse_value))

                details['reverse_value'] = reverse_value
                if cards.filter(card_number=reverse_value).exists():
                    card = cards.filter(card_number=reverse_value).first()
                    details['card_name'] = " - " + card.person.get_name() + ", " + card.card_name
                
                person_ids = cards.values_list('person_id', flat=True).distinct()

                persons_objects = persons_objects.filter(id__in=person_ids)

            else:
                persons_objects = persons_objects.annotate(
                        full_name=Concat('first_name', Value(' '), 'last_name')
                    ).filter(Q(full_name__icontains=value) | Q(note__icontains=value))

                
        persons_objects.order_by('last_name', 'first_name')

    persons = [person.get_details() for person in persons_objects]
    details.update({"persons": persons, "is_admin": is_admin,"allow_feedback": allow_feedback})
    if end_index > 0:
        details['current_index'] = end_index
    return render(request, 'locks/persons_list.html', details)


def hostings_ichilov(request):
    allow_feedback = request.user.is_superuser or request.user.allow_feedback
    hostings_objects = HostingIchilov.objects.all().order_by('-date_add')
    hostings = [hosting.get_details() for hosting in hostings_objects]
    return render(request, 'locks/hostings_ichilov.html', {"hostings": hostings, "allow_feedback": allow_feedback})

def person(request, person_id=None):
    allow_feedback = request.user.is_superuser or request.user.allow_feedback
    user = request.user
    is_admin = request.user.is_superuser
    data = {"allow_feedback": allow_feedback, "is_admin": is_admin, "user": user}
    if person_id:
        person = Person.objects.get(id=person_id)
        person_details = person.get_details()
        data['details'] = person_details
        data['type_view'] = 'edit'
    else:
        data['new'] = 'edit'

    
    return render(request, f'locks/person.html', data)



def add_person(request):
    return render(request, 'locks/add_person.html')


def merge_person(request, source_person_id):
    data = {}
    source_person = Person.objects.get(id=source_person_id)
    data['source_person_id'] = source_person.id
    data['source_person_name'] = source_person.last_name + " " + source_person.first_name
    persons = Person.get_persons(True)
    data['persons'] = [person for person in persons if person['id'] != source_person_id and len(person['name']) > 2]

    return render(request, 'locks/merge_person.html', data)

def do_merge_person(request, source_person_id, target_person_id):
    Card.objects.filter(person_id=source_person_id).update(person_id=target_person_id)
    Phone.objects.filter(person_id=source_person_id).update(person_id=target_person_id)
    Hosting.objects.filter(guest_id=source_person_id).update(guest_id=target_person_id)
    Hosting.objects.filter(patient_id=source_person_id).update(patient_id=target_person_id)
    HostingIchilov.objects.filter(person_id=source_person_id).update(person_id=target_person_id)
    HostingRecord.objects.filter(guest_id=source_person_id).update(guest_id=target_person_id)

    target_person = Person.objects.get(id=target_person_id)
    LockRecord.objects.filter(person_id=source_person_id).update(person_id=target_person_id, person_name=target_person.get_name())

    Person.objects.get(id=source_person_id).delete()
    
    return JsonResponse({'success': True})

def get_person_child_list(request, type_child, person_id):
    person = Person.objects.get(id=person_id)
    child_list = person.get_child_list(type_child, request.user)

    return JsonResponse({'success': True, 'child_list': child_list})

def save_person(request):
    try:
        person_data = json.loads(request.body.decode('utf-8'))

        if "person_id" in person_data:
            person_id = person_data.get("person_id")
            obj = Person.objects.get(id=person_id)
            last_obj = obj
            create = False
        else:
            id_number = person_data.get('id_number')
            obj, create = Person.objects.get_or_create(id_number=id_number)
        
        # טיפול מיוחד בתאריך לידה
        birth_date = person_data.pop('birth_date', None)
        
        for key, value in person_data.items():
            setattr(obj, key, value)
        
        # שמירת תאריך לידה אם קיים
        if birth_date:
            from datetime import datetime
            try:
                birth_date_obj = datetime.strptime(birth_date, '%Y-%m-%d').date()
                obj.birth_date = birth_date_obj
            except (ValueError, AttributeError):
                pass  # אם יש שגיאה בפורמט, לא שומרים
        
        obj.not_to_view = False

        from .translite_ai import transliterate_hebrew_name
        if len(obj.first_name) > 0:  
            obj.first_name_eng = transliterate_hebrew_name(obj.first_name)
        if len(obj.last_name) > 0 :
            obj.last_name_eng = transliterate_hebrew_name(obj.last_name)

        obj.save()
        person_id = obj.id
    

    except IntegrityError as e:
        return JsonResponse({"success": False, "error": "השם שהוזן כבר קיים במערכת"}, status=500)

    return JsonResponse({"success": True, "person_id": person_id})



def delete_person(request, person_id):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    permissions_objects = Person.objects.get(id=person_id).get_permissions()
    if not request.user.is_superuser:
        lock_ids_with_permission = LockUser.objects.filter(user=request.user).values_list('lock_id', flat=True)
        permissions_objects = permissions_objects.filter(~Q(lock_id__in=lock_ids_with_permission)).exists()
        if permissions_objects:
            return JsonResponse({'success': False, 'error': 'אין אפשרות למחוק את האורח מכיון שיש לו הרשאות למנעולים אחרים'}, status=200)

    Person.objects.get(id=person_id).delete_person(user_name)
    locks_to_sync = [permission.lock.id for permission in permissions_objects]
    return JsonResponse({"success": True, "locks_to_sync": locks_to_sync})


def get_role_options(request):
    roles = Person.get_all_roles()
    roles = [{"label": role, "value": role} for role in roles]
    return JsonResponse({"roles": roles})

def check_id_number(request, id_number):
    data = {}
    if Person.objects.filter(id_number=id_number).exists():
        data['details'] = Person.objects.get(id_number=id_number).get_details()
        data['exist'] = True
    else:
        data['exist'] = False
    return JsonResponse(data)


    

def get_person_permission_objects(request, person_id):
    person = Person.objects.get(id=person_id)
    cards_objects = person.card_set.all()
    cards = [card.get_details() for card in cards_objects]
    phones_objects = person.phone_set.all()
    phones = [phone.get_details() for phone in phones_objects]
    return JsonResponse({"cards": cards, "phones": phones})

def records(request):
    seven_days_ago_timestamp_ms = get_seven_days_ago_timestamp_ms()

    if not request.user.is_superuser:
        lock_ids_with_permission = LockUser.objects.filter(user=request.user).values_list('lock_id', flat=True)
        records_objects = LockRecord.objects.filter(lock_date__gte=seven_days_ago_timestamp_ms, lock_id__in=lock_ids_with_permission).order_by("-lock_date")[:500]
    else:
        records_objects = LockRecord.objects.filter(lock_date__gte=seven_days_ago_timestamp_ms).order_by("-lock_date")[:500]

    records = [record.get_details() for record in records_objects]
    
    # return JsonResponse({"records": records})
    return render(request, 'locks/records.html', {"records": records})


def phone(request, phone_id):
    phone = Phone.objects.get(id=phone_id)
    phone_details = phone.get_details()
    phone_details['phone_id'] = phone_id

    return render(request, 'locks/phone.html', phone_details)

def add_phone(request, person_id):
    data = {}
    data['person_id'] = person_id
    return render(request, 'locks/phone.html')

def save_phone(request):
    phone_data = json.loads(request.body.decode('utf-8'))
    phone_id=phone_data.get('phone_id')
    phone_data.pop('phone_id')
    try:
        if phone_id == '':
            new_phone = Phone(**phone_data)
            new_phone.save()
            phone_id = new_phone.id
        else:
            Phone.objects.filter(id=phone_id).update(**phone_data)

    except IntegrityError as e:
        return JsonResponse({"success": False, "error": "המספר או השם שהוזנו כבר קיימים במערכת"}) 

    return JsonResponse({"success": True, "phone_id": phone_id})

def main_sync(request):
    do_main_sync()
    return JsonResponse({"success": True})



def sync(request, lock_id):
    lock = Lock.objects.get(id=lock_id)
    lock.sync_permissions()
    lock.update_details()
    
    send_message_to_browser('refresh_locks')
    return JsonResponse({"success": True})


def import_locks_records(request):
    do_import_locks_records()
    return JsonResponse({"success": True})

def do_import_locks_records():
    for lock in Lock.objects.all():
        lock.import_records()
    send_message_to_browser('refresh_records')

def transmission(request):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    locks = request.POST.getlist('locks[]')
    do_transmission(user_name, locks)
    return JsonResponse({"success": True})

def transmission_all(request):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    do_transmission(user_name, "all")
    return JsonResponse({"success": True})

def do_transmission(user_name, locks):
    send_message_to_browser('start_transmissions')
    cleen_empty_transmission()
    if locks == 'all':
        locks = [lock.id for lock in Lock.objects.all()]

    for lock_id in locks:
        lock = Lock.objects.get(id=lock_id)
        transmission = lock.transmission_permissions()
        if transmission:
            lock.sync_permissions()
    
    cleen_record_to_delete(user_name)
    send_message_to_browser('end_transmissions')

def cleen_empty_transmission():
    seven_days_ago_timestamp_ms = get_seven_days_ago_timestamp_ms()
    Transmission.objects.filter(last_transmission__lte=seven_days_ago_timestamp_ms).delete()

    all_ids_permissions = Permission.objects.all().values_list('id', flat=True)
    active_permissions = Permission.objects.filter(status_record=0).values_list('id', flat=True)
    
    Transmission.objects.filter(
        ~Q(status_transmission=2) & 
        (~Q(permission_id__in=all_ids_permissions) | 
        Q(permission_id__in=active_permissions))
    ).delete()

def resend_transmission(request, transmission_id):
    transmission = Transmission.objects.get(id=transmission_id)
    permission_id = transmission.permission_id
    permission = Permission.objects.get(id=permission_id)
    send_message_to_browser('start_transmissions')
    permission.transmission()
    send_message_to_browser('end_transmissions')
    return JsonResponse({"success": True})

def delete_transmission(request, transmission_id):
    transmission = Transmission.objects.get(id=transmission_id)
    permission_id = transmission.permission_id
    permission = Permission.objects.get(id=permission_id)
    if transmission.type_action == 1:    
        permission.delete()
    elif transmission.type_action == 2:
        permission.status_record = 0
        permission.save()
        
    transmission.delete()
    send_message_to_browser('refresh_permissions_screens')
    send_amount_in_transmission()
    return JsonResponse({"success": True})


def resend_all_transmissions(request):
    transmissions = get_tranmissions(request.user)
    transmissions.filter(status_transmission__in=[3,4]).update(status_transmission=0, last_transmission=None, amount_attempts=0)

    return JsonResponse({"success": True})


def get_cards_list(request, person_id=None):
    cards = Card.get_cards(person_id)
    return JsonResponse({'success': True, "cards": cards}, status=200)


def delete_person_child(request, type_child, object_id):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    if type_child == 'permissions':
        permissions_object = Permission.objects.get(id=object_id)
        permissions_object.delete_permission(user_name)
        locks_to_sync = [permissions_object.lock.id]

    elif type_child == 'cards':
        card_object = Card.objects.get(id=object_id)
        permissions_objects = Permission.objects.filter(object_id=object_id, type_object_id=8)
        if not request.user.is_superuser:
            lock_ids_with_permission = LockUser.objects.filter(user=request.user).values_list('lock_id', flat=True)
            permissions_objects = permissions_objects.filter(~Q(lock_id__in=lock_ids_with_permission)).exists()
            if permissions_objects:
                return JsonResponse({'success': False, 'error': 'אין אפשרות למחוק את הכרטיס מכיון שהוא מכיל הרשאות למנעולים אחרים'}, status=200)
        
        card_object.delete_card(user_name)
        
        locks_to_sync = [permission.lock.id for permission in permissions_objects]

    elif type_child == 'phones':
        phone_object = Phone.objects.get(id=object_id)

        permissions_objects = Permission.objects.filter(object_id=object_id, type_object_id=16)
        if not request.user.is_superuser:
            lock_ids_with_permission = LockUser.objects.filter(user=request.user).values_list('lock_id', flat=True)
            permissions_objects = permissions_objects.filter(~Q(lock_id__in=lock_ids_with_permission)).exists()
            if permissions_objects:
                return JsonResponse({'success': False, 'error': 'אין אפשרות למחוק את הפלאפון מכיון שהוא מכיל הרשאות למנעולים אחרים'}, status=200)

        phone_object.delete()
        locks_to_sync = []

    return JsonResponse({'success': True, "locks_to_sync": locks_to_sync}, status=200)


def delete_cards(request):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    data = json.loads(request.body.decode('utf-8'))
    phones = data.get("phones", [])
    cards = data.get("cards", [])

    all_permissions = Permission.objects.filter(Q(type_object_id=8, object_id__in=cards) | Q(type_object=16, object_id__in=phones))

    if not request.user.is_superuser:
        lock_ids_with_permission = LockUser.objects.filter(user=request.user).values_list('lock_id', flat=True)
        permissions_not_in_user_locks = all_permissions.filter(~Q(lock_id__in=lock_ids_with_permission)).exists()
        if permissions_not_in_user_locks:
            return JsonResponse({'success': False, 'error': "לא ניתן למחוק את אמצעי הזיהוי מכיון שהם מכילים הרשאות למנעולים אחרים"}, status=200)


    cards_permissions = all_permissions.filter(type_object_id=8)
    locks_to_sync = [permission.lock.id for permission in cards_permissions]

    for card in cards:
        card_objects = Card.objects.get(id=card)
        card_objects.delete_card(user_name)
    
    Phone.objects.filter(id__in=phones).delete()

    return JsonResponse({'success': True, 'locks_to_sync': locks_to_sync}, status=200)

def card(request, card_id):
    card = Card.objects.get(id=card_id)
    card_details = card.get_details()
    
    persons_cards_objects = Card.objects.filter(person_id=card.person_id).exclude(id=card.id) 
    persons_cards = [{"id": card.id, "name": card.card_name} for card in persons_cards_objects]

    card_details['card_id'] = card_id
    card_details['person_id'] = card.person.id
    card_details['person_name'] = card.person.get_name()
    card_details['persons_cards'] = persons_cards
    persons = Person.get_persons(True)
    card_details['persons'] = persons
    return render(request, 'locks/card.html', card_details)

def add_card(request, person_id):
    card_details = {}

    persons_cards_objects = Card.objects.filter(person_id=person_id) 
    persons_cards = [{"id": card.id, "name": card.card_name} for card in persons_cards_objects]
    count_cards = persons_cards_objects.count()

    card_details['person_id'] = person_id
    card_details['persons_cards'] = persons_cards
    card_details['person_name'] = Person.objects.get(id=person_id).get_name()
    card_details['card_id'] = ''
    if count_cards == 0:
        card_details['card_name'] = 'קבוע'
    else:
        card_details['card_name'] = "כרטיס " + str(count_cards + 1)
    persons = Person.get_persons(True)
    card_details['persons'] = persons
    return render(request, 'locks/card.html', card_details)

def reverse_card_number(card_number):
    card_number = int(card_number)
    hex_string = hex(card_number)[2:]
    hex_string = '0' * (len(hex_string) % 2) + hex_string
    reversed_hex_string = ''.join(reversed([hex_string[i:i+2] for i in range(0, len(hex_string), 2)]))
    reversed_number = int(reversed_hex_string, 16)
    
    return reversed_number

def save_card(request):
    data = json.loads(request.body.decode('utf-8'))
    card_id=data.get('card_id')
    person_name = data.get('person_name')
    reverse_card = data.get('reverse_card')
    card_data = data.get('card_data')
    import_permission_from = data.get('import_permission_from')

    if reverse_card == 1:
        card_data['card_number'] = reverse_card_number(card_data.get('card_number'))

    locks_to_sync = []

    try:
        if card_id == '':
            new_card = Card(**card_data)
            new_card.save()
            card_id = new_card.id
        else:

            exist_card = Card.objects.get(id=card_id)
            for key, value in card_data.items():
                setattr(exist_card, key, value)
            exist_card.save()

            exist_card.rename_care()

            permissions = Permission.objects.filter(type_object_id=8, object_id=exist_card.id)
            locks_to_sync = [permission.lock_id for permission in permissions]

    except ValidationError as e:
        errors = e.message_dict
        message = next(iter(errors.values()))[0]
        if "does not exist" in message:
            message = "יש לבחור אורח" 
        if message == 'Card with this Card name and Person already exists.':
            message = "לאורח " + person_name + " כבר קיים כרטיס בשם " + card_data.get('card_name')
        if message == "Card with this Card number already exists.":
            message = "הכרטיס כבר קיים במערכת"
        return JsonResponse({"success": False, "error": message}, status=500)

    except IntegrityError as e:
        message = "לאורח " + person_name + " כבר קיים כרטיס בשם " + card_data.get('card_name')
        return JsonResponse({"success": False, "error": message}, status=500)
    
    locks_to_sync = []
    if import_permission_from:
        permission_to_duplicate = Permission.objects.filter(object_id=import_permission_from, type_object_id=8)
        locks_to_sync += [permission.lock_id for permission in permission_to_duplicate]
        for permission in permission_to_duplicate:
            new_permission = Permission(
                object_id=card_id,
                type_object_id=8,
                lock=permission.lock,
                type_permission=permission.type_permission,
                start_date = permission.start_date,
                end_date = permission.end_date,
                cyclic_config = permission.cyclic_config,
                grant_by = permission.grant_by,
                status_record=1
            )
            new_permission.save()

            transmission_record = Transmission()
            transmission_record.type_action = 1
            transmission_record.lock_name = new_permission.lock.lock_alias
            transmission_record.card_number = new_permission.content_object.card_number
            transmission_record.card_description = new_permission.content_object.get_full_description()
            transmission_record.permission_id = new_permission.id
            transmission_record.user_transmission = f"{request.user.first_name} {request.user.last_name}"
            transmission_record.save()

    return JsonResponse({"success": True, "card_id": card_id, "locks_to_sync": locks_to_sync})

def delete_card(request, card_id):
    user_name = f"{request.user.first_name} {request.user.last_name}"
    Card.objects.get(id=card_id).delete_card(user_name)
    return JsonResponse({"success": True})

def change_coffee_card_status(request, card_id, status):
    card = Card.objects.get(id=card_id)
    card.coffee_card = status
    card.save()
    return JsonResponse({"success": True})

def get_excel_cards(request):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Font
    
    data = json.loads(request.body.decode('utf-8'))
    cards = data.get("cards", [])

    cards_objects = Card.objects.filter(id__in=cards)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cards"
    ws.sheet_view.rightToLeft = True

    headers = ['שם', 'תפקיד', 'פלאפון', 'מספר כרטיס', 'שם כרטיס','קפה', 'תאריך הוספה'] 
    ws.append(headers)

    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gray color
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = gray_fill
        cell.font = bold_font

    for card in cards_objects:
        ws.append([card.person.get_name(), card.person.role, card.person.person_phone, card.card_number, card.card_name, card.coffee_card, time_to_datetime(card.date_add)])  
    
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cards.xlsx'
    
    wb.save(response)
    return response

def get_excel_hostings_summary(request):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Font
    from datetime import datetime
    
    data = json.loads(request.POST.get('data', '{}'))
    filters = data.get("filters", {})
    
    # Apply filters
    q_object = Q()
    for key, value in filters.items():
        if key.startswith('*'):
            q_object &= Hosting.convert_filter(key[1:], value)
        else:
            q_object &= Q(**{key: value})
    
    hostings = Hosting.objects.filter(q_object).select_related('house', 'guest', 'patient').order_by('lodging_start')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "אירוחים מסכם"
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ['דירה', 'שם אורח', 'ת.ז. מלווה', 'טלפון אורח', 'עיר אורח', 'סטטוס אורח', 
               'שם מטופל', 'ת.ז. מטופל', 'טלפון מטופל', 'עיר מטופל', 'מחלקה', 'גורם מפנה',
               'תאריך כניסה', 'תאריך יציאה', 'לילות', 'אנשים בבית', 'הערות']
    ws.append(headers)
    
    # Styling
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    bold_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = gray_fill
        cell.font = bold_font
    
    # Data rows
    for hosting in hostings:
        guest_name = hosting.guest.get_name()
        guest_phone = hosting.guest.person_phone or ''
        guest_city = hosting.guest.city or ''
        guest_id_number = hosting.guest.id_number or ''
        
        if hosting.guest_is_patient:
            status_guest = "מטופל"
            patient_name = ''
            patient_id_number = guest_id_number
            patient_phone = ''
            patient_city = ''
            guest_id_number = ''  # מלווה ריק כי האורח הוא המטופל
        else:
            status_guest = "מלווה"
            if hosting.affinity:
                status_guest += f' ({hosting.affinity})'
            patient_name = hosting.patient.get_name() if hosting.patient else ''
            patient_id_number = hosting.patient.id_number if hosting.patient else ''
            patient_phone = hosting.patient.person_phone if hosting.patient else ''
            patient_city = hosting.patient.city if hosting.patient else ''
        
        nights = (hosting.lodging_end - hosting.lodging_start).days
        
        row = [
            hosting.house.description,
            guest_name,
            guest_id_number,
            guest_phone,
            guest_city,
            status_guest,
            patient_name,
            patient_id_number,
            patient_phone,
            patient_city,
            hosting.hospital_ward or '',
            hosting.trigger or '',
            hosting.lodging_start,
            hosting.lodging_end,
            nights,
            hosting.persons_in_house,
            hosting.note or ''
        ]
        ws.append(row)
    
    # Borders
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    today = datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename=אירוחים_מסכם_{today}.xlsx'
    
    wb.save(response)
    return response

def get_excel_hostings_daily(request):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Font
    from datetime import datetime, timedelta
    
    data = json.loads(request.POST.get('data', '{}'))
    filters = data.get("filters", {})
    
    # Apply filters
    q_object = Q()
    for key, value in filters.items():
        if key.startswith('*'):
            q_object &= Hosting.convert_filter(key[1:], value)
        else:
            q_object &= Q(**{key: value})
    
    hostings = Hosting.objects.filter(q_object).select_related('house', 'guest', 'patient').order_by('lodging_start')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "אירוחים יומי"
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ['תאריך', 'דירה', 'שם אורח', 'ת.ז. מלווה', 'טלפון אורח', 'עיר אורח', 'סטטוס אורח', 
               'שם מטופל', 'ת.ז. מטופל', 'טלפון מטופל', 'עיר מטופל', 'מחלקה', 'גורם מפנה',
               'תאריך כניסה', 'תאריך יציאה', 'לילות', 'אנשים בבית', 'הערות']
    ws.append(headers)
    
    # Styling
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    bold_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = gray_fill
        cell.font = bold_font
    
    # Data rows - iterate through each day
    for hosting in hostings:
        guest_name = hosting.guest.get_name()
        guest_phone = hosting.guest.person_phone or ''
        guest_city = hosting.guest.city or ''
        guest_id_number = hosting.guest.id_number or ''
        
        if hosting.guest_is_patient:
            status_guest = "מטופל"
            patient_name = ''
            patient_id_number = guest_id_number
            patient_phone = ''
            patient_city = ''
            guest_id_number = ''  # מלווה ריק כי האורח הוא המטופל
        else:
            status_guest = "מלווה"
            if hosting.affinity:
                status_guest += f' ({hosting.affinity})'
            patient_name = hosting.patient.get_name() if hosting.patient else ''
            patient_id_number = hosting.patient.id_number if hosting.patient else ''
            patient_phone = hosting.patient.person_phone if hosting.patient else ''
            patient_city = hosting.patient.city if hosting.patient else ''
        
        nights = (hosting.lodging_end - hosting.lodging_start).days
        
        # Iterate through each day in the hosting period
        current_date = hosting.lodging_start
        while current_date <= hosting.lodging_end:
            row = [
                current_date,
                hosting.house.description,
                guest_name,
                guest_id_number,
                guest_phone,
                guest_city,
                status_guest,
                patient_name,
                patient_id_number,
                patient_phone,
                patient_city,
                hosting.hospital_ward or '',
                hosting.trigger or '',
                hosting.lodging_start,
                hosting.lodging_end,
                nights,
                hosting.persons_in_house,
                hosting.note or ''
            ]
            ws.append(row)
            current_date += timedelta(days=1)
    
    # Borders
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    today = datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename=אירוחים_יומי_{today}.xlsx'
    
    wb.save(response)
    return response

from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def view_unlock_by_phone(request):
    body_str = request.body.decode('utf-8')
    data_dict = parse_qs(body_str)
    data_dict = {key: value[0] for key, value in data_dict.items()}
    answer = unlock_by_phone_menu(data_dict)

    return HttpResponse(answer, content_type='text/plain')

def send_feedback(request, type_hosting, hosting_id):
    from urllib.parse import urlencode
    import requests

    if type_hosting == 'tel':
        hosting = Hosting.objects.get(id=hosting_id)
        person = hosting.guest
    else:
        hosting = HostingIchilov.objects.get(id=hosting_id)
        person = hosting.person
    
    if len(person.email) == 0:
       return JsonResponse({"success": False, "error": "אין אימייל מוגדר"}) 

    hosting.send_feedback = 1
    hosting.save()

    form_data = {
        'object_id': '2720',
        'data': f'{{"c1": "{person.first_name}", "c2": "{person.last_name}", "c3": "ampSoftware1@gmail.com", "c5": "{person.person_phone}"}}'
    }

    url = 'https://forms.sogomatic.com/r-lev/wp-json/wp/v2/submit_flow'

    headers = {
        'sogo-token': 'e39af0322d161a8f4824a8944cfd028abf2f526e66b121fb826bbd43c9f600d8', 
    }

    response = requests.post(url, data=form_data, headers=headers)
    return JsonResponse({"success": True, "response": response.text})

def guest_form(request):
    
    if request.method == 'POST':
        try:
            id_number = request.POST.get('IdNumber')
            first_name = request.POST.get('firstName')
            last_name = request.POST.get('lastName')
            cell_phone = request.POST.get('cellPhone')
            email = request.POST.get('email')
            street = request.POST.get('street')
            house_number = request.POST.get('houseNumber')
            city = request.POST.get('city')
            type_reg = request.POST.get('type_reg')
            note = request.POST.get('note')
            
            if not id_number or not first_name or not last_name:
                return JsonResponse({
                    'success': False,
                    'error': 'יש למלא את כל השדות החובה'
                }, status=400)
            
            if type_reg == '1' or type_reg == '2':
                role = 'אורח'
            elif type_reg == '3':
                role = 'צוות'
            elif type_reg == '4':
                role = 'נותן שירות'
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'יש לבחור סוג רישום'
                }, status=400)

            person, created = Person.objects.get_or_create(id_number=id_number)
            
            from .translite_ai import transliterate_hebrew_name, translate_hospital_department

            person.first_name = first_name
            person.last_name = last_name
            person.first_name_eng = transliterate_hebrew_name(first_name)
            person.last_name_eng = transliterate_hebrew_name(last_name)
            person.person_phone = cell_phone
            person.email = email
            person.address = street
            person.house_number = house_number
            person.city = city
            person.role = role
            person.note = note

            person.save()

            if type_reg == '1':
                
                type_guest = request.POST.get('type_guest')

                hosting_record = HostingRecord(guest=person)
                
                if type_guest == '1':
                    hosting_record.guest_is_patient = 1
                    hospital_ward = request.POST.get('GuestHospitalWard')
                else:
                    patiant_id_number = request.POST.get('patiantIdNumber')
                    patiant_first_name = request.POST.get('patiantfirstName')
                    patiant_last_name = request.POST.get('patiantlastName')
                    affinity = request.POST.get('affinity')
                    hospital_ward = request.POST.get('hospitalWard')

                    patient_person_obj, created = Person.objects.get_or_create(id_number=patiant_id_number)
                    patient_person_obj.first_name = patiant_first_name
                    patient_person_obj.last_name = patiant_last_name
                    patient_person_obj.first_name_eng = transliterate_hebrew_name(patiant_first_name)
                    patient_person_obj.last_name_eng = transliterate_hebrew_name(patiant_last_name)

                    if created:
                        patient_person_obj.role = "אורח"
                    patient_person_obj.save()

                    hosting_record.guest_is_patient = 0
                    hosting_record.patient = patient_person_obj
                    hosting_record.affinity = affinity
                
                hosting_record.hospital_ward = hospital_ward
                hosting_record.save()
                
                # קביעת המטופל - אם האורח הוא המטופל או אם יש מטופל נפרד
                patient_person = person if type_guest == '1' else hosting_record.patient
                
                # טיפול בשמירת תאריך לידה של המטופל
                patient_birth_date = request.POST.get('patientBirthDate')
                if patient_birth_date and patient_person:
                    from datetime import datetime
                    try:
                        # המרת התאריך מפורמט YYYY-MM-DD לאובייקט date
                        birth_date_obj = datetime.strptime(patient_birth_date, '%Y-%m-%d').date()
                        patient_person.birth_date = birth_date_obj
                        patient_person.save()
                    except ValueError:
                        pass  # אם יש שגיאה בפורמט, לא שומרים
                
                # טיפול בשמירת קובץ תעודת זהות
                patient_id_file = request.FILES.get('patientIdFile')
                if patient_id_file and patient_person:
                    # שמירת הקובץ באמצעות Django FileField
                    patient_person.id_number_file = patient_id_file
                    patient_person.save()

            try:
                webhooks = WebHooks()
                webhooks.send_sms("התקבל טופס חדש מ" + first_name + " " + last_name)
            except:
                pass  # אם יש שגיאה בשליחת SMS, ממשיכים בכל זאת

            # החזרת תגובת JSON במקום redirect
            return JsonResponse({
                'success': True,
                'message': 'הטופס נשלח בהצלחה'
            })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'אירעה שגיאה בשמירת הטופס: {str(e)}'
            }, status=500)

    # תצוגת הטופס
    return render(request, 'locks/guest_form.html', {})

def guest_form_hostings_ichilov(request):
    from .translite_ai import transliterate_hebrew_name, translate_hospital_department
    
    if request.method == 'GET':
        id_number = request.GET.get('IdNumber', None)
        first_name = request.GET.get('firstName', None)
        last_name = request.GET.get('lastName', None)
        person_phone = request.GET.get('Phone', None )
        email = request.GET.get('email', None)
        street = request.GET.get('street', None)
        house_number = request.GET.get('houseNumber', None)
        city = request.GET.get('city', None)
        
        person, created = Person.objects.get_or_create(id_number=id_number)
        
        person.first_name = first_name
        person.last_name = last_name
        person.first_name_eng = transliterate_hebrew_name(first_name)
        person.last_name_eng = transliterate_hebrew_name(last_name)
        person.person_phone = person_phone
        person.email = email
        person.address = street
        person.house_number = house_number
        person.city = city
        person.role = "אורח"
        person.not_to_view = 1
        person.save()


        hospital_ward = request.GET.get('hospitalWard', None)
        hospital_ward_eng = translate_hospital_department(hospital_ward)
        lodging_start = request.GET.get('lodgingStart', None)
        lodging_start = datetime.strptime(lodging_start, "%d-%m-%Y").date()

        lodging_end= request.GET.get('lodgingEnd', None)
        lodging_end = datetime.strptime(lodging_end, "%d-%m-%Y").date()

        hosting = HostingIchilov(              
            person=person,
            hospital_ward=hospital_ward,
            hospital_ward_eng=hospital_ward_eng,
            lodging_start=lodging_start,
            lodging_end=lodging_end
        )
        hosting.save()

        return JsonResponse({"status": 'OK'})

def house_page(request, house_code):
    house_object = get_object_or_404(House, link_code=house_code)
    house = house_object.get_details()
    

    current_date = datetime.now().date()
    thre_months_ago = current_date - timedelta(days=180)
    hostings_objects = house_object.hosting_set.annotate(
        is_still_active=Case(When(lodging_end__gt=localdate(), then=Value(True)),  
        default=Value(False), 
        output_field=BooleanField(),
    )).filter(lodging_start__lt=current_date, lodging_start__gte=thre_months_ago).order_by('-is_still_active', '-lodging_start')

    hostings = [hosting.get_details_for_page() for hosting in hostings_objects]
    data = {"house": house, "hostings": hostings}
    # return JsonResponse(data)

    webhooks = WebHooks()
    webhooks.send("דף תורם", 1, "", "", "", house['description'])

    return render(request, 'locks/house_page.html', data)


def donor_page(request, donor_code):
    donor_object = get_object_or_404(Donor, code=donor_code)
    current_date = datetime.now().date()
    six_months_ago = current_date - timedelta(days=180)

    hostings_objects = Hosting.objects.annotate(
        is_still_active=Case(When(lodging_end__gt=localdate(), then=Value(True)),  
        default=Value(False), 
        output_field=BooleanField(),
    )).filter(lodging_start__lt=current_date, lodging_start__gte=six_months_ago).order_by('-is_still_active', '-lodging_start')

    hostings = [hosting.get_details_for_donor_page() for hosting in hostings_objects]
    data = {"donor_name": donor_object.name, "hostings": hostings}
    # return JsonResponse(data)

    webhooks = WebHooks()
    webhooks.send("דף תורם", 2, "", "", "", donor_object.name)

    return render(request, 'locks/donor_hosting_page.html', data)

def transmissions(request):
    return render(request, 'locks/transmissions.html')

def get_transmissions_list(request):
    from django.db.models import Case, When, Value, IntegerField
    transmissions = get_tranmissions(request.user)
    seven_days_ago_timestamp_ms = get_seven_days_ago_timestamp_ms()
    transmissions_objects = transmissions.filter(~Q(status_transmission=2) | Q(last_transmission__gte=seven_days_ago_timestamp_ms)) \
        .annotate(has_last_transmission=Case(When(last_transmission__isnull=True, then=Value(0)),
                                                  default=Value(1),
                                                  output_field=IntegerField())) \
        .order_by("has_last_transmission", "-last_transmission")
    transmissions = [transmission.get_details() for transmission in transmissions_objects]
    return JsonResponse({"transmissions": transmissions})

def get_details_user(user):

    if user.last_login is None:
        last_login = '-'
    else:
        last_login = user.last_login
        server_timezone = datetime.now(timezone.utc).astimezone().tzinfo
        dt_object_local = last_login.replace(tzinfo=pytz.utc).astimezone(server_timezone)
        last_login = dt_object_local.strftime("%Y-%m-%d %H:%M:%S")

    user_details = {
        "user_id": user.id,
        "username": user.username,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "email": user.email,
        "is_superuser": user.is_superuser,
        "last_login":  last_login,
        "allow_main": user.allow_main,
        "allow_coffee_cards": user.allow_coffee_cards,
        "allow_passage_modes": user.allow_passage_modes,
        "allow_hostings_ichilov": user.allow_hostings_ichilov,
        "allow_feedback": user.allow_feedback,
        "allow_houses": user.allow_houses,
        "allow_locks_records": user.allow_locks_records,
        "allow_permissions": user.allow_permissions,
        "allow_hostings": user.allow_hostings,
    }
    return user_details

def users(request):
    users = User.objects.all()
    users_details = [get_details_user(user) for user in users]
    return render(request, 'locks/users.html', {"users": users_details})

def add_user(request):
    return render(request, 'locks/user.html')

def user(request, user_id):
    user_details = get_details_user(User.objects.get(id=user_id))
    
    return render(request, 'locks/user.html', user_details)

def reset_password(request, user_id, new_password):

    user_object = User.objects.get(pk=user_id)
    user_object.set_password(new_password)
    user.save()
                
    return JsonResponse({"success": True})

def save_user(request):
    user_data = json.loads(request.body.decode('utf-8'))
    user_id = user_data.get('user_id')
    if user_id == '':
        user_data.pop('user_id')
        try:
            new_user = User.objects.create_user( **user_data)
        except IntegrityError as e:
            if "Duplicate entry" in str(e):
                return JsonResponse({"success": False, "error": "שם משתמש כבר קיים"})
            else:
                return JsonResponse({"success": False, "error": str(e)})
        
    else:
        user_object = User.objects.get(id=user_id)
        change_pssword = False
        if "password" in user_data:
            from django.contrib.auth.hashers import make_password
            change_pssword = True
            password = user_data.get('password')
            user_data.pop('password')
       
        try:
            for key, value in user_data.items():
                setattr(user_object, key, value)

            if change_pssword:
                user_object.password = make_password(password)
            user_object.save()

        except IntegrityError as e:
            if "Duplicate entry" in str(e):
                return JsonResponse({"success": False, "error": "שם משתמש כבר קיים"})
            else:
                return JsonResponse({"success": False, "error": str(e)})
        
    return JsonResponse({"success": True, "user_id": user_id})

def delete_user(request, user_id):
    User.objects.get(id=user_id).delete()
    return JsonResponse({"success": True})

def upload_person_id_file(request):
    if request.method == 'POST':
        file = request.FILES.get('id_number_file')
        person_id = request.POST.get('person_id')
        
        if not file or not person_id:
            return JsonResponse({'success': False, 'error': 'missing_data'}, status=400)
        
        try:
            import uuid
            import os
            from django.core.files.base import ContentFile
            
            person = Person.objects.get(id=person_id)
            
            # מחיקת הקובץ הישן אם קיים
            if person.id_number_file:
                person.id_number_file.delete(save=False)
            
            # יצירת שם קובץ חדש בלי תווים עבריים
            file_extension = os.path.splitext(file.name)[1]
            new_filename = f"id_{person.id}_{uuid.uuid4().hex[:8]}{file_extension}"
            
            # שמירת הקובץ החדש עם השם החדש
            person.id_number_file.save(new_filename, ContentFile(file.read()), save=True)
            
            # החזרת רק הנתיב היחסי
            file_path = str(person.id_number_file.name)
            response_data = {
                'success': True, 
                'file_path': file_path
            }
            return HttpResponse(json.dumps(response_data), content_type='application/json', status=200)
        except Person.DoesNotExist:
            return HttpResponse(json.dumps({'success': False, 'error': 'not_found'}, ensure_ascii=False), content_type='application/json; charset=utf-8', status=404)
        except Exception as e:
            return HttpResponse(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False), content_type='application/json; charset=utf-8', status=500)
    
    return HttpResponse(json.dumps({'success': False, 'error': 'method_not_allowed'}, ensure_ascii=False), content_type='application/json; charset=utf-8', status=405)

def delete_person_id_file(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            person_id = data.get('person_id')
            
            if not person_id:
                return JsonResponse({'success': False, 'error': 'missing_person_id'}, status=400)
            
            person = Person.objects.get(id=person_id)
            
            # מחיקת הקובץ אם קיים
            if person.id_number_file:
                person.id_number_file.delete(save=False)
                person.id_number_file = None
                person.save()
                
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'no_file_exists'}, status=400)
                
        except Person.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'person_not_found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'method_not_allowed'}, status=405)

def upload_file(request):
    file = request.FILES.get('file')
        
    if file:
        random_filename = f"{uuid.uuid4()}{os.path.splitext(file.name)[1]}"
        file_path = os.path.join(settings.BASE_DIR, 'media/files', random_filename)

        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return JsonResponse({'success': True, 'file_path': file_path}, status=200)

def get_form_links(reqeust):
    import requests
    import re

    flow_name = reqeust.GET.get('flow_name')
    
    url = "https://forms.sogomatic.com/r-lev/wp-json/wp/v2/get-flow-rows?flow_name=" + flow_name
    headers = {
        "sogo-token": "e39af0322d161a8f4824a8944cfd028abf2f526e66b121fb826bbd43c9f600d8"
    }

    response = requests.get(url, headers=headers)

    data = response.json() 
    data = data['data']
    links = []

    for row in data:
        if "params" in row['fields']:
            match = re.search(r'params";s:115:"([^"]+)"', row['fields'])
            if match:
                long_code = match.group(1)
                form = {}
                form['name'] = row['name']
                form['updated_at'] = row['updated_at']
                form['link'] = "https://forms.sogomatic.com/r-lev/blog/pdf/vacation-registration/?params=" + long_code
                form['status'] = row['status'] 
                links.append(form)
    return JsonResponse(links, safe=False)
    
def test(request):
    from .translite_ai import translate_hospital_department
    hostings = Hosting.objects.all()
    for hosting in hostings:
        if len(hosting.hospital_ward) > 0 and not hosting.hospital_ward_eng:  
            hosting.hospital_ward_eng = translate_hospital_department(hosting.hospital_ward)
        
        hosting.save()

